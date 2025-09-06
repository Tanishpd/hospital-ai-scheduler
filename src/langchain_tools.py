"""
LangChain Tools for Medical Appointment Scheduling Agent
Provides specialized tools for database operations, calendar management, and email integration.
"""

from langchain_core.tools import tool  # type: ignore
from typing import Dict, List, Optional, Any
import pandas as pd  # type: ignore
import re
from datetime import datetime, timedelta
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
from dotenv import load_dotenv  # type: ignore

load_dotenv()

# Global data store for the session
patients_df = None
appointments_df = None

def load_databases():
    """Load patient and appointment databases"""
    global patients_df, appointments_df
    try:
        patients_df = pd.read_csv('data/patients.csv')
    except FileNotFoundError:
        patients_df = pd.DataFrame(columns=[
            # Original fields (preserved for backward compatibility)
            'patient_id', 'name', 'dob', 'doctor', 'last_visit', 
            'phone', 'email', 'insurance_carrier', 'member_id', 'group_id',
            # New demographic fields
            'first_name', 'last_name', 'middle_name', 'gender', 'ssn',
            # Address information
            'street_address', 'city', 'state', 'zip_code', 'alt_phone',
            # Enhanced insurance information
            'policy_holder_name', 'relationship_to_patient',
            # Emergency contact
            'emergency_contact_name', 'emergency_contact_relationship', 'emergency_contact_phone',
            # Medical information
            'primary_concern', 'medical_history', 'current_medications', 'allergies', 'family_history',
            # Lifestyle information
            'smoking_status', 'alcohol_consumption', 'exercise_frequency',
            # System fields
            'created_date', 'updated_date'
        ])
    
    try:
        appointments_df = pd.read_csv('data/appointments.csv')
    except FileNotFoundError:
        appointments_df = pd.DataFrame(columns=[
            'appointment_id', 'patient_name', 'doctor', 'date', 'time', 
            'duration', 'type', 'status', 'notes'
        ])

@tool
def lookup_patient_in_emr(name: str, date_of_birth: str) -> Dict[str, Any]:
    """
    Look up patient information in Electronic Medical Records (EMR) system.
    
    Args:
        name: Patient's full name
        date_of_birth: Patient's date of birth in YYYY-MM-DD format
        
    Returns:
        Dictionary containing patient information or None if not found
    """
    if patients_df is None:
        load_databases()
    
    if patients_df is None or patients_df.empty:
        return {
            'found': False,
            'message': 'Patient database is not available'
        }
    
    name_clean = name.strip().lower()
    
    # Exact match search
    exact_match = patients_df[
        (patients_df['name'].str.lower() == name_clean) & 
        (patients_df['dob'] == date_of_birth)
    ]
    
    if not exact_match.empty:
        patient_data = exact_match.iloc[0].to_dict()
        patient_data['found'] = True
        patient_data['classification'] = 'returning_patient'
        return patient_data
    
    # Search for partial name match
    name_parts = name_clean.split()
    if len(name_parts) >= 2:
        first_name = name_parts[0]
        last_name = name_parts[-1]
        
        partial_match = patients_df[
            (patients_df['name'].str.lower().str.contains(first_name)) |
            (patients_df['name'].str.lower().str.contains(last_name))
        ]
        
        if not partial_match.empty:
            patient_data = partial_match.iloc[0].to_dict()
            patient_data['found'] = True
            patient_data['classification'] = 'returning_patient'
            patient_data['note'] = 'Partial name match found'
            return patient_data
    
    return {
        'found': False,
        'classification': 'new_patient',
        'name': name,
        'dob': date_of_birth,
        'note': 'New patient - not found in EMR system'
    }

@tool
def get_comprehensive_patient_info(patient_id: str) -> Dict[str, Any]:
    """
    Get comprehensive patient information including all intake form fields.
    
    Args:
        patient_id: Patient's unique identifier
        
    Returns:
        Dictionary containing comprehensive patient information
    """
    if patients_df is None:
        load_databases()
    
    if patients_df is None or patients_df.empty:
        return {
            'found': False,
            'message': 'Patient database is not available'
        }
    
    # Find patient by ID
    patient_match = patients_df[patients_df['patient_id'] == int(patient_id)]
    
    if patient_match.empty:
        return {
            'found': False,
            'message': f'Patient with ID {patient_id} not found'
        }
    
    patient_data = patient_match.iloc[0].to_dict()
    
    # Organize data into logical sections
    comprehensive_info = {
        'found': True,
        'basic_info': {
            'patient_id': patient_data.get('patient_id'),
            'first_name': patient_data.get('first_name'),
            'last_name': patient_data.get('last_name'),
            'middle_name': patient_data.get('middle_name'),
            'full_name': patient_data.get('name'),
            'date_of_birth': patient_data.get('dob'),
            'gender': patient_data.get('gender'),
            'ssn': patient_data.get('ssn')
        },
        'contact_info': {
            'phone': patient_data.get('phone'),
            'alt_phone': patient_data.get('alt_phone'),
            'email': patient_data.get('email'),
            'address': {
                'street': patient_data.get('street_address'),
                'city': patient_data.get('city'),
                'state': patient_data.get('state'),
                'zip_code': patient_data.get('zip_code')
            }
        },
        'insurance_info': {
            'carrier': patient_data.get('insurance_carrier'),
            'member_id': patient_data.get('member_id'),
            'group_id': patient_data.get('group_id'),
            'policy_holder': patient_data.get('policy_holder_name'),
            'relationship': patient_data.get('relationship_to_patient')
        },
        'emergency_contact': {
            'name': patient_data.get('emergency_contact_name'),
            'relationship': patient_data.get('emergency_contact_relationship'),
            'phone': patient_data.get('emergency_contact_phone')
        },
        'medical_info': {
            'primary_concern': patient_data.get('primary_concern'),
            'medical_history': patient_data.get('medical_history'),
            'current_medications': patient_data.get('current_medications'),
            'allergies': patient_data.get('allergies'),
            'family_history': patient_data.get('family_history')
        },
        'lifestyle': {
            'smoking_status': patient_data.get('smoking_status'),
            'alcohol_consumption': patient_data.get('alcohol_consumption'),
            'exercise_frequency': patient_data.get('exercise_frequency')
        },
        'care_info': {
            'assigned_doctor': patient_data.get('doctor'),
            'last_visit': patient_data.get('last_visit'),
            'created_date': patient_data.get('created_date'),
            'updated_date': patient_data.get('updated_date')
        }
    }
    
    return comprehensive_info

@tool
def search_patients_by_criteria(search_term: str = "", field: str = "name") -> List[Dict[str, Any]]:
    """
    Search for patients using various criteria and comprehensive patient information.
    
    Args:
        search_term: Search term to look for
        field: Field to search in (name, phone, email, doctor, insurance_carrier, etc.)
        
    Returns:
        List of matching patients with their comprehensive information
    """
    if patients_df is None:
        load_databases()
    
    if patients_df is None or patients_df.empty:
        return [{
            'found': False,
            'message': 'Patient database is not available'
        }]
    
    search_term_lower = search_term.lower().strip()
    
    # Define searchable fields
    valid_fields = [
        'name', 'first_name', 'last_name', 'phone', 'email', 'doctor',
        'insurance_carrier', 'city', 'state', 'gender', 'primary_concern',
        'medical_history', 'allergies', 'current_medications'
    ]
    
    if field not in valid_fields:
        return [{
            'found': False,
            'message': f'Invalid search field. Valid fields: {", ".join(valid_fields)}'
        }]
    
    # Perform search
    if field in patients_df.columns:
        # Handle null values
        mask = patients_df[field].notna() & patients_df[field].astype(str).str.lower().str.contains(search_term_lower, na=False)
        matching_patients = patients_df[mask]
    else:
        return [{
            'found': False,
            'message': f'Field "{field}" not found in database'
        }]
    
    if matching_patients.empty:
        return [{
            'found': False,
            'message': f'No patients found matching "{search_term}" in field "{field}"'
        }]
    
    # Return comprehensive info for all matching patients
    results = []
    for _, patient in matching_patients.iterrows():
        # Organize patient data into logical sections manually
        patient_data = patient.to_dict()
        comprehensive_info = {
            'found': True,
            'basic_info': {
                'patient_id': patient_data.get('patient_id'),
                'first_name': patient_data.get('first_name'),
                'last_name': patient_data.get('last_name'),
                'middle_name': patient_data.get('middle_name'),
                'full_name': patient_data.get('name'),
                'date_of_birth': patient_data.get('dob'),
                'gender': patient_data.get('gender'),
                'ssn': patient_data.get('ssn')
            },
            'contact_info': {
                'phone': patient_data.get('phone'),
                'alt_phone': patient_data.get('alt_phone'),
                'email': patient_data.get('email'),
                'address': {
                    'street': patient_data.get('street_address'),
                    'city': patient_data.get('city'),
                    'state': patient_data.get('state'),
                    'zip_code': patient_data.get('zip_code')
                }
            },
            'care_info': {
                'assigned_doctor': patient_data.get('doctor'),
                'last_visit': patient_data.get('last_visit')
            }
        }
        results.append(comprehensive_info)
    
    return results

@tool
def get_doctor_availability(doctor_name: Optional[str] = None, preferred_date: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Get available appointment slots for doctors.
    
    Args:
        doctor_name: Specific doctor to check (optional)
        preferred_date: Preferred date in YYYY-MM-DD format (optional)
        
    Returns:
        List of available appointment slots
    """
    # Load doctor schedules from Excel files
    import os
    from openpyxl import load_workbook  # type: ignore
    
    schedule_dir = 'data/doctor_schedules'
    available_slots = []
    
    if not os.path.exists(schedule_dir):
        return [{
            'error': 'Doctor schedule files not found',
            'note': 'Please ensure doctor schedule Excel files are available'
        }]
    
    # Get current date for filtering
    current_date = datetime.now().date()
    target_date = None
    if preferred_date:
        try:
            target_date = datetime.strptime(preferred_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    for filename in os.listdir(schedule_dir):
        if filename.endswith('.xlsx'):
            doctor = filename.replace('_schedule.xlsx', '').replace('_', ' ').title()
            
            # Skip if specific doctor requested and this isn't it
            if doctor_name and doctor.lower() != doctor_name.lower():
                continue
                
            try:
                wb = load_workbook(os.path.join(schedule_dir, filename))
                ws = wb.active
                
                if ws is None:
                    continue
                
                # Skip header row
                for row in ws.iter_rows(min_row=2, values_only=True):  # type: ignore
                    if len(row) >= 5 and row[0] and row[1]:  # Date and Time columns
                        slot_date = row[0]
                        slot_time = row[1]
                        specialty = row[2] if row[2] else 'General Medicine'
                        duration = row[3] if row[3] else 30
                        status = row[4] if row[4] else 'Available'
                        
                        # Convert date if it's a datetime object
                        if hasattr(slot_date, 'date') and callable(getattr(slot_date, 'date')):
                            slot_date = slot_date.date()  # type: ignore
                        elif isinstance(slot_date, str):
                            try:
                                slot_date = datetime.strptime(slot_date, '%Y-%m-%d').date()
                            except ValueError:
                                continue
                        else:
                            continue  # Skip invalid date formats
                        
                        # Filter by date criteria
                        if slot_date < current_date:
                            continue
                        if target_date and slot_date != target_date:
                            continue
                        if not isinstance(status, str) or status.lower() != 'available':  # type: ignore
                            continue
                            
                        available_slots.append({
                            'doctor': doctor,
                            'specialty': specialty,
                            'date': slot_date.strftime('%Y-%m-%d'),
                            'time': str(slot_time),
                            'duration_minutes': duration,
                            'status': status
                        })
                        
            except Exception as e:
                continue
    
    # Sort by date and time
    available_slots.sort(key=lambda x: (x['date'], x['time']))
    
    # Return up to 10 slots to avoid overwhelming the LLM
    return available_slots[:10] if available_slots else [{
        'note': 'No available slots found',
        'suggestion': 'Try a different date or doctor'
    }]

@tool
def book_appointment(patient_name: str, doctor: str, date: str, time: str, appointment_type: str, duration_minutes: int = 30) -> Dict[str, Any]:
    """
    Book an appointment for a patient.
    
    Args:
        patient_name: Patient's full name
        doctor: Doctor's name
        date: Appointment date in YYYY-MM-DD format
        time: Appointment time in HH:MM format
        appointment_type: Type of appointment (new_patient, follow_up, etc.)
        duration_minutes: Duration in minutes (default 30)
        
    Returns:
        Booking confirmation details
    """
    global appointments_df
    if appointments_df is None:
        load_databases()
    
    # Generate appointment ID
    appointment_count = len(appointments_df) if appointments_df is not None else 0
    appointment_id = f"APT-{datetime.now().strftime('%Y%m%d')}-{appointment_count + 1:04d}"
    
    # Create appointment record
    new_appointment = {
        'appointment_id': appointment_id,
        'patient_name': patient_name,
        'doctor': doctor,
        'date': date,
        'time': time,
        'duration': duration_minutes,
        'type': appointment_type,
        'status': 'scheduled',
        'notes': f'Booked via AI agent on {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    }
    
    # Add to dataframe
    appointments_df = pd.concat([appointments_df, pd.DataFrame([new_appointment])], ignore_index=True)
    
    # Save to CSV
    try:
        appointments_df.to_csv('data/appointments.csv', index=False)
    except Exception as e:
        return {
            'success': False,
            'error': f'Failed to save appointment: {str(e)}'
        }
    
    return {
        'success': True,
        'appointment_id': appointment_id,
        'patient_name': patient_name,
        'doctor': doctor,
        'date': date,
        'time': time,
        'duration_minutes': duration_minutes,
        'type': appointment_type,
        'confirmation': f'Appointment {appointment_id} successfully booked for {patient_name} with Dr. {doctor} on {date} at {time}'
    }

@tool
def send_appointment_confirmation(patient_email: str, appointment_details: Dict[str, Any]) -> Dict[str, Any]:
    """
    Send appointment confirmation email to patient.
    
    Args:
        patient_email: Patient's email address
        appointment_details: Dictionary containing appointment information
        
    Returns:
        Email sending status
    """
    try:
        # Email configuration
        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        email_user = os.getenv('EMAIL_USER')
        email_password = os.getenv('EMAIL_PASSWORD')
        
        if not email_user or not email_password:
            return {
                'success': False,
                'error': 'Email credentials not configured'
            }
        
        # Create email content
        subject = f"Appointment Confirmation - {appointment_details.get('appointment_id', 'N/A')}"
        
        body = f"""
        Dear {appointment_details.get('patient_name', 'Patient')},

        Your appointment has been successfully scheduled!

        Appointment Details:
        - Appointment ID: {appointment_details.get('appointment_id', 'N/A')}
        - Doctor: Dr. {appointment_details.get('doctor', 'N/A')}
        - Date: {appointment_details.get('date', 'N/A')}
        - Time: {appointment_details.get('time', 'N/A')}
        - Duration: {appointment_details.get('duration_minutes', 30)} minutes
        - Type: {appointment_details.get('type', 'N/A')}

        Please arrive 15 minutes early for check-in.

        If you need to reschedule or cancel, please contact our office at least 24 hours in advance.

        Thank you for choosing our medical practice!

        Best regards,
        Medical Scheduling System
        """
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = email_user
        msg['To'] = patient_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        
        # Send email
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(email_user, email_password)
        text = msg.as_string()
        server.sendmail(email_user, patient_email, text)
        server.quit()
        
        return {
            'success': True,
            'message': f'Confirmation email sent to {patient_email}',
            'appointment_id': appointment_details.get('appointment_id', 'N/A')
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f'Failed to send email: {str(e)}'
        }

@tool
def validate_patient_information(name: Optional[str] = None, phone: Optional[str] = None, email: Optional[str] = None, date_of_birth: Optional[str] = None) -> Dict[str, Any]:
    """
    Validate patient information format and completeness.
    
    Args:
        name: Patient's full name
        phone: Phone number
        email: Email address
        date_of_birth: Date of birth in YYYY-MM-DD format
        
    Returns:
        Validation results with any errors or corrections
    """
    validation_results = {
        'valid': True,
        'errors': [],
        'corrections': {}
    }
    
    # Validate name
    if name:
        if not re.match(r'^[a-zA-Z\s]{2,50}$', name.strip()):
            validation_results['valid'] = False
            validation_results['errors'].append('Name must contain only letters and spaces (2-50 characters)')
        else:
            validation_results['corrections']['name'] = name.strip().title()
    
    # Validate phone
    if phone:
        digits = re.sub(r'\D', '', phone)
        if len(digits) != 10:
            validation_results['valid'] = False
            validation_results['errors'].append('Phone number must be 10 digits')
        else:
            validation_results['corrections']['phone'] = f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    
    # Validate email
    if email:
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email.lower()):
            validation_results['valid'] = False
            validation_results['errors'].append('Invalid email address format')
        else:
            validation_results['corrections']['email'] = email.lower()
    
    # Validate date of birth
    if date_of_birth:
        try:
            birth_date = datetime.strptime(date_of_birth, '%Y-%m-%d')
            if birth_date.date() > datetime.now().date():
                validation_results['valid'] = False
                validation_results['errors'].append('Date of birth cannot be in the future')
            elif birth_date.date() < datetime(1900, 1, 1).date():
                validation_results['valid'] = False
                validation_results['errors'].append('Date of birth seems too old')
            else:
                validation_results['corrections']['date_of_birth'] = birth_date.strftime('%Y-%m-%d')
        except ValueError:
            validation_results['valid'] = False
            validation_results['errors'].append('Date of birth must be in YYYY-MM-DD format')
    
    return validation_results

@tool
def get_insurance_verification_status(insurance_carrier: str, member_id: str, group_id: Optional[str] = None) -> Dict[str, Any]:
    """
    Verify insurance information and coverage status.
    
    Args:
        insurance_carrier: Insurance company name
        member_id: Insurance member ID
        group_id: Insurance group ID (optional)
        
    Returns:
        Insurance verification status
    """
    # Simulate insurance verification
    common_carriers = [
        'Blue Cross Blue Shield', 'Aetna', 'Cigna', 'United Healthcare',
        'Kaiser Permanente', 'Humana', 'Anthem', 'Medicare', 'Medicaid'
    ]
    
    # Basic validation
    if not insurance_carrier or not member_id:
        return {
            'verified': False,
            'status': 'incomplete_information',
            'message': 'Insurance carrier and member ID are required'
        }
    
    # Simulate verification based on realistic scenarios
    carrier_lower = insurance_carrier.lower()
    is_known_carrier = any(carrier.lower() in carrier_lower for carrier in common_carriers)
    
    if is_known_carrier and len(member_id) >= 8:
        return {
            'verified': True,
            'status': 'active',
            'carrier': insurance_carrier,
            'member_id': member_id,
            'group_id': group_id,
            'copay': '$25',
            'deductible_remaining': '$150',
            'coverage_type': 'Standard',
            'message': 'Insurance verified successfully'
        }
    else:
        return {
            'verified': False,
            'status': 'verification_required',
            'carrier': insurance_carrier,
            'member_id': member_id,
            'message': 'Manual verification required - please contact insurance department'
        }

# Initialize databases on import
load_databases()
